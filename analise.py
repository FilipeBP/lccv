import numpy as np
import pandas as pd
from entrada import Planeta
from openpyxl import load_workbook

class Integral():
    """
    Classe para usar métodos de integração
    e exportar os dados calculados para o mesmo arquivo original.
    """
    def __init__(self,list,tf,dt):
        #Recebendo a lista de planetas.
        self.planets = list

        #Variáveis para os cálculos.
        self.tf = tf
        self.dt = dt
        self.n = round(tf/dt) + 1
        self.G = -6.674e-2 #-11 #G deve ser negativo para o cálculo ocorrer corretamente.

        #Criando matriz de dimensões de acordo com o número de planetas e a quantidade de cálculos irá ser feitos.
        self.x = np.zeros((len(list),self.n))
        self.y = np.zeros((len(list),self.n))
        self.z = np.zeros((len(list),self.n))
        self.vx = np.zeros((len(list),self.n))
        self.vy = np.zeros((len(list),self.n))
        self.vz = np.zeros((len(list),self.n))

        #Atribuindo os valores iniciais para o primeiro elemento da matriz, de acordo com seu planeta.
        for i in range(len(list)):
            self.x[i,0] = list[i].x0
            self.y[i,0] = list[i].y0
            self.z[i,0] = list[i].z0
            self.vx[i,0] = list[i].v0x
            self.vy[i,0] = list[i].v0y
            self.vz[i,0] = list[i].v0z

        #Checando se o planeta 1 será fixo.
        if list[0].x0 == 0 and list[0].y0 == 0 and list[0].z0 == 0 and list[0].v0x == 0 and list[0].v0y == 0 and list[0].v0z == 0:
            self.start = 1
        else:
            self.start = 0

    def getCollision(self,model,matrix):
        """Método para receber os modelos de colisão e a matriz de colisão."""
        self.model = model
        self.matrix = matrix

    def F(self,var,i,j):
        """Método para calcular a aceleração de acordo com a força gravitacional do planeta, e uma possível colisão."""
        f = 0
        for k in range(len(self.planets)):
            if k != i:
                #Calculando a força gravitacional.
                r = abs((self.x[k,j-1]-self.x[i,j-1])**2 + (self.y[k,j-1]-self.y[i,j-1])**2 + (self.z[k,j-1]-self.z[i,j-1])**2)**(1/2)
                f -= self.G*self.planets[k].mass*(var[k,j-1]-var[i,j-1])/r**3

                #Possível colisão.
                if r < self.planets[i].radius + self.planets[k].radius:
                    if k > i:
                        model = self.matrix[i,k]
                    else:
                        model = self.matrix[k,i]

                    K = self.model[model-1,1]
                    modulo_dx = r - self.planets[i].radius - self.planets[k].radius
                    dx = modulo_dx*(var[k,j-1]-var[i,j-1])/r

                    f += K*dx
        return f

    def newton(self):
        """Método para calcular os dados usando o método de Newton"""
        for j in range(1,self.n):
            for i in range(self.start,len(self.planets)):
                #Atualizando a aceleração para cada variável.
                ax = self.F(self.x,i,j)
                ay = self.F(self.y,i,j)
                az = self.F(self.z,i,j)

                #Atualizando a velocidade para cada variável
                self.vx[i,j] = self.vx[i,j-1] +self.dt*ax
                self.vy[i,j] = self.vy[i,j-1] +self.dt*ay
                self.vz[i,j] = self.vz[i,j-1] +self.dt*az

                #Atualizando a posição para cada variável
                self.x[i,j] = self.x[i,j-1] + self.dt*self.vx[i,j]
                self.y[i,j] = self.y[i,j-1] + self.dt*self.vy[i,j]
                self.z[i,j] = self.z[i,j-1] + self.dt*self.vz[i,j]

    def verlet(self):
        """Método para calcular os dados usando o método de Verlet"""
        for j in range(1,self.n):
            for i in range(self.start,len(self.planets)):
                #Atualizando a aceleração para cada variável.
                ax = self.F(self.x,i,j)
                ay = self.F(self.y,i,j)
                az = self.F(self.z,i,j)

                #A primeira velocidade calculada é onde ocorre o deslocamento do tempo para a velocidade para verlet.
                if j == 1:
                    self.vx[i,j] = self.vx[i,j-1] +self.dt/2*ax
                    self.vy[i,j] = self.vy[i,j-1] +self.dt/2*ay
                    self.vz[i,j] = self.vz[i,j-1] +self.dt/2*az
                else:
                    self.vx[i,j] = self.vx[i,j-1] +self.dt*ax
                    self.vy[i,j] = self.vy[i,j-1] +self.dt*ay
                    self.vz[i,j] = self.vz[i,j-1] +self.dt*az

                #Atualizando a posição para cada variável
                self.x[i,j] = self.x[i,j-1] + self.dt*self.vx[i,j]
                self.y[i,j] = self.y[i,j-1] + self.dt*self.vy[i,j]
                self.z[i,j] = self.z[i,j-1] + self.dt*self.vz[i,j]

    def exportToFile(self,file,passo):
        """Método para exportar os dados calculados para o mesmo arquivo de entrada"""
        n = round(self.tf/(passo*self.dt))
        t = [i*n*self.dt for i in range(passo+1)]

        #Criando listas de acordo com o passo para o exportar ao arquivo.
        x_export = [[self.x[i,j] for j in range(self.n) if j%n==0] for i in range(len(self.planets))]
        y_export = [[self.y[i,j] for j in range(self.n) if j%n==0] for i in range(len(self.planets))]
        z_export = [[self.z[i,j] for j in range(self.n) if j%n==0] for i in range(len(self.planets))]
        vx_export = [[self.vx[i,j] for j in range(self.n) if j%n==0] for i in range(len(self.planets))]
        vy_export = [[self.vy[i,j] for j in range(self.n) if j%n==0] for i in range(len(self.planets))]
        vz_export = [[self.vz[i,j] for j in range(self.n) if j%n==0] for i in range(len(self.planets))]

        #Lendo o arquivo com o método load_workbook do openpyxl para adicionar worksheets ao arquivo.
        excelfile = pd.ExcelFile(file)
        book = load_workbook(file)
        writer = pd.ExcelWriter(file, engine='openpyxl')
        writer.book = book

        #Exportando os dados para cada planeta.
        for i in range(len(self.planets)):
            df = pd.DataFrame({'t':t,'x':x_export[i],'y':y_export[i],'z':z_export[i],'vx':vx_export[i],'vy':vy_export[i],'vz':vz_export[i]})
            if f'Planeta {i+1}' in excelfile.sheet_names:
                continue
            df.to_excel(writer,sheet_name=f'Planeta {i+1}')

        writer.save()
        writer.close()
